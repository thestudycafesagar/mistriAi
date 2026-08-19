import time
import re
import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =================================================================
# Keyword matching helper (regex, word-boundary-anchored)
# =================================================================

def keyword_matches(text, keyword):
    """
    True if `keyword` appears in `text` (already lowercased) with a
    non-alphanumeric boundary (string start/end, space, "/", "(", "-", etc.)
    on AT LEAST ONE side — i.e. not fully embedded inside a longer, unrelated
    word on BOTH sides. Plain substring checks (`keyword in text`) miss this:
    "date" is a genuine substring of "Updated" (embedded on both sides — a
    false positive) but ALSO of "TransactionDate" (a legitimate suffix, with
    a real column-header meaning — a boundary at the string's end). Likewise
    "ref" is a substring of "Preferred" (embedded both sides — false
    positive) and of "Cheque/Reference No" (right after "/" — legitimate).
    Requiring a boundary on only one side (not `\\bkeyword\\b` on both)
    rejects the "Updated"/"Preferred" cases while still matching "date" as a
    suffix (TransactionDate/ValueDate/PostingDate — all real header
    patterns) and "withdraw" as a PREFIX of "withdrawals"/"withdraws" (a
    real column-header variant this parser already relies on recognizing,
    confirmed on a real Canara statement).
    """
    esc = re.escape(keyword)
    return re.search(r'(?<![a-z0-9])' + esc + r'|' + esc + r'(?![a-z0-9])', text) is not None


def any_keyword_matches(text, keywords):
    return any(keyword_matches(text, kw) for kw in keywords)

# Date separators are usually the ASCII "/" or "-", but some statement
# generators (confirmed on a real Cosmos Co-operative Bank statement) render
# dates using the Unicode MINUS SIGN (U+2212, "−") instead of the ASCII
# hyphen (U+002D, "-") — visually near-identical, but a distinct codepoint
# that a plain "-" in a character class does NOT match. Every date-pattern
# regex in this file used to require plain "/"/"-", so on that statement
# every date silently failed to match at all — the anchor-row detection in
# extract_transactions_by_word_coords() found zero anchors and bailed
# entirely (mirrors the earlier UCO month-name DATE_PAT fix: a narrower
# pattern than the real document uses causes a silent, total failure, not a
# partial one). Also includes the common hyphen/dash lookalikes (U+2010–
# U+2015) defensively, on the same reasoning as the Unicode minus sign.
DATE_SEP = r'[/\-‐-―−]'

# =================================================================
# EXISTING: Robust Helper functions (Untouched)
# =================================================================

def merge_multirow_header(rows):
    num_cols = max(len(r) for r in rows) if rows else 0
    merged = []
    for col_idx in range(num_cols):
        parts = []
        for row in rows:
            if col_idx < len(row):
                val = str(row[col_idx]).strip() if row[col_idx] else ''
                if val:
                    parts.append(val)
        merged.append(' '.join(parts) if parts else f'Column_{col_idx + 1}')
    return merged

def collapse_table_to_row(table):
    if not table:
        return []
    num_cols = max(len(r) for r in table)
    result = [''] * num_cols
    for row in table:
        for col_idx, cell in enumerate(row):
            if col_idx < num_cols:
                val = str(cell).strip() if cell else ''
                if val:
                    result[col_idx] = (result[col_idx] + ' ' + val).strip() if result[col_idx] else val
    return result

def get_explicit_columns(page, target_headers):
    words = page.extract_words()
    rows = {}
    for w in words:
        y = round(w['top'], 1)
        matched_y = None
        for key_y in rows:
            if abs(key_y - y) < 4:
                matched_y = key_y
                break
        if matched_y is None:
            matched_y = y
            rows[matched_y] = []
        rows[matched_y].append(w)
        
    best_match_count = 0
    best_v_lines = None
    
    for y, rwords in rows.items():
        text = ' '.join(w['text'].lower() for w in rwords)
        matches = sum(1 for t in target_headers if t in text)
        has_date = 'date' in text or 'txn' in text or 'transaction' in text
        
        if has_date and matches >= 2 and matches > best_match_count:
            best_match_count = matches
            v_lines = [page.bbox[0]]
            for w in rwords:
                w_text = w['text'].lower()
                if any(t in w_text or w_text in t for t in target_headers):
                    v_lines.append(w['x0'] - 2)
            v_lines.append(page.bbox[2])
            best_v_lines = sorted(list(set(v_lines)))

    return best_v_lines


def extract_transactions_by_word_coords(pdf_path, target_headers):
    """
    Word-coordinate-based extractor for PDFs where:
    - The header spans multiple visual y-lines (e.g. 'Transaction\\nDate' split across rows)
    - Narration text wraps to adjacent lines above or below the date row

    Algorithm:
      1. Detect all header y-rows (within 20px of first matching row) → derive col x-starts
      2. Pass 1: collect anchor rows = rows with a date at the leftmost column x-zone
      3. Pass 2: assign continuation rows to nearest anchor by y-distance
    """
    # The month group accepts either 1-2 digits (01-04-2026) OR a 3+ letter
    # month name/abbreviation (01-Apr-2026, 01/April/2026) — a purely numeric
    # pattern matches zero rows on statements that print the month as text
    # (confirmed on a real UCO Bank statement), which silently zeroes out
    # anchor_indices entirely, makes this whole extractor bail with (None,
    # None), and falls back to the far cruder "standard style" table loop —
    # producing exactly the kind of narration-shifted, rows-glued-to-the-
    # wrong-transaction corruption this extractor exists to avoid.
    DATE_PAT = re.compile(r'^\d{1,2}' + DATE_SEP + r'(?:\d{1,2}|[A-Za-z]{3,9})' + DATE_SEP + r'\d{2,4}(?:\s|$)', re.IGNORECASE)

    def band_words_excluding_date_rows(words, center_y, start_offset=-5, end_offset=25):
        """
        Collects every word within [center_y+start_offset, center_y+end_offset],
        grouped into visual rows, excluding any row whose own leftmost word
        looks like a real transaction date. A genuine multi-line header
        continuation (e.g. "Transaction" wrapping to "Date" on the next line)
        never starts with a date value — only a real transaction row does.
        Without this guard, a transaction row that sits close below the
        header (confirmed on a real MZRB/SBI statement: only ~20px below,
        well inside the generous +25px multi-line-header allowance) gets
        swept into the header band, and any keyword its own narration
        coincidentally contains (e.g. "Deposit" inside "By Cash Deposit by
        self") gets mistaken for a header label — merging it into an
        adjacent column's name and corrupting it.
        """
        band = [w for w in words if center_y + start_offset <= w['top'] <= center_y + end_offset]
        y_rows = {}
        for w in band:
            wy = round(w['top'])
            matched = next((ky for ky in y_rows if abs(ky - wy) <= 5), None)
            if matched is None:
                matched = wy
                y_rows[matched] = []
            y_rows[matched].append(w)
        result = []
        for wy, row_words in y_rows.items():
            row_words_sorted = sorted(row_words, key=lambda rw: rw['x0'])
            if DATE_PAT.match(row_words_sorted[0]['text'].strip()):
                continue
            result.extend(row_words)
        return result

    with pdfplumber.open(pdf_path) as pdf:
        all_page_words = []
        page_heights = []
        for page in pdf.pages:
            all_page_words.append(page.extract_words())
            page_heights.append(page.bbox[3])

    if not all_page_words:
        return None, None

    # ── Step 1: Detect header rows ────────────────────────────────────────
    # Find the first y-row with 'date'/'txn' + ≥2 target-header keyword matches.
    # Then collect all rows within 25px below it (multi-line header support).
    #
    # Bank-statement exports commonly repeat the SAME title/letterhead/column
    # header block at the top of EVERY page (confirmed on a real Cosmos
    # Co-operative Bank statement, whose page 2 repeats the bank's name/
    # address, account number, statement period, and the column header row
    # itself). The column-position derivation below only needs to run once
    # (columns don't move page to page), but the SKIP must happen on every
    # page — otherwise a later page's repeated header/title text is treated
    # as ordinary "continuation" text and gets merged straight into that
    # page's first real transaction (its nearest anchor by y-distance),
    # corrupting the narration and gluing header words like "Particulars"/
    # "Withdrawals" into the wrong columns. Confirmed: this silently dropped
    # a real ₹2,000 withdrawal's amount into a garbled merged row, making the
    # extracted total ₹2,000 short of the statement's own printed Grand
    # Total. Detect every page's own header row independently (not just the
    # first), so Step 2 can skip each page's header band on that page too.
    header_y_range = None  # (y_start, y_end)
    data_start_page = 0
    first_header_y = None
    page_header_end_y = {}  # pg_idx -> y below which real transaction rows start

    for pg_idx, words in enumerate(all_page_words):
        y_rows = {}
        for w in words:
            y = round(w['top'])
            matched = next((ky for ky in y_rows if abs(ky - y) <= 5), None)
            if matched is None:
                matched = y
                y_rows[matched] = []
            y_rows[matched].append(w)

        for y in sorted(y_rows):
            rw = sorted(y_rows[y], key=lambda w: w['x0'])
            rt = ' '.join(w['text'].lower() for w in rw)
            matches = sum(1 for t in target_headers if t in rt)
            if ('date' in rt or 'txn' in rt) and matches >= 2:
                if first_header_y is None:
                    first_header_y = y
                    data_start_page = pg_idx
                # Precise end-of-header-band for THIS page: collect words in
                # a generous band below this row (handles a multi-line
                # header), then take the max y among only the words that
                # actually match a target keyword. A flat "+25" offset here
                # is too coarse — some statements pack transaction rows only
                # ~8px apart (confirmed on a real Cosmos statement's page 2),
                # so a flat +25 swallowed the first TWO real transactions
                # along with the header, dropping a real ₹2,000 withdrawal
                # and a real ₹4,900 deposit. This mirrors the precise
                # calculation already used for data_start_page below, so
                # every page's skip is equally tight, not a blunt guess.
                band_words = band_words_excluding_date_rows(words, y)
                matched_ys = [
                    w['top'] for w in band_words
                    if any(t in w['text'].lower().strip('():') or w['text'].lower().strip('():') in t for t in target_headers)
                ]
                page_header_end_y[pg_idx] = (max(matched_ys) + 5) if matched_ys else (y + 10)
                break

    if first_header_y is None:
        return None, None

    # Collect all words in the header y-band (first_header_y ± 25px),
    # excluding any row that's actually a nearby real transaction (see
    # band_words_excluding_date_rows for why — confirmed on a real MZRB/SBI
    # statement, where the first transaction's own narration "By Cash
    # Deposit by self" sat only ~20px below the header and had its "Deposit"
    # word swept in, merging "Particulars" + "Deposit" into one bogus
    # column name).
    words_pg0 = all_page_words[data_start_page]
    header_band_words = band_words_excluding_date_rows(words_pg0, first_header_y)

    # Derive column x-starts from header band words that match target keywords
    col_entries = []   # list of (x0, name, x1)
    for w in sorted(header_band_words, key=lambda w: w['x0']):
        wt = w['text'].lower().strip('():')
        if any(t in wt or wt in t for t in target_headers):
            # Merge with previous col if gap between words is small
            if col_entries and (w['x0'] - col_entries[-1][2]) < 10:
                col_entries[-1] = (col_entries[-1][0], col_entries[-1][1] + ' ' + w['text'], w['x1'])
            else:
                col_entries.append((w['x0'], w['text'], w['x1']))

    if len(col_entries) < 3:
        return None, None

    col_names = [c[1] for c in col_entries]
    col_starts = sorted([c[0] for c in col_entries])
    # page_header_end_y[data_start_page] was already computed precisely by
    # the per-page loop above (same "max matched-keyword-word y + 5" logic).

    def get_col_idx(x):
        """Assign x to nearest column by x0 boundaries."""
        idx = 0
        for i, cs in enumerate(col_starts):
            if x >= cs - 8:
                idx = i
        return idx

    # Date column x-zone = everything to the left of the second column start
    date_col_x_max = col_starts[1] - 5 if len(col_starts) > 1 else 120

    # ── Amount-column clustering (fixes right-aligned numeric columns) ────
    # Header labels are left-anchored, but amounts (WITHDRAWALS/DEPOSITS/BALANCE
    # etc.) are right-aligned. A wide value (e.g. "7,07,196.00Dr") can start well
    # to the LEFT of its own header's x0, causing plain x0-vs-header-start
    # matching to snap it into the PREVIOUS column instead. Since right-aligned
    # values in the same column always end (x1) at nearly the same x, cluster the
    # x1 of every amount-looking token across the whole document and match
    # clusters (left→right) to the amount-keyword headers (left→right).
    AMOUNT_KEYWORDS = ('balance', 'amount', 'withdraw', 'deposit', 'credit', 'debit')
    AMOUNT_VALUE_RE = re.compile(r'^[₹$€£]?\s*-?\(?[\d,]+\.\d{1,2}\)?\s*(dr|cr)?$', re.IGNORECASE)

    numeric_cols_sorted = sorted(
        (col_starts[i], i) for i, name in enumerate(col_names)
        if any(k in name.lower() for k in AMOUNT_KEYWORDS)
    )

    cluster_to_col = {}
    amount_clusters = []
    if len(numeric_cols_sorted) >= 2:
        amount_x1s = sorted(
            w['x1'] for words in all_page_words for w in words
            if AMOUNT_VALUE_RE.match(w['text'].strip())
        )
        if amount_x1s:
            clusters = []
            cur = [amount_x1s[0]]
            for x in amount_x1s[1:]:
                if x - cur[-1] <= 40:
                    cur.append(x)
                else:
                    clusters.append(sum(cur) / len(cur))
                    cur = [x]
            clusters.append(sum(cur) / len(cur))
            # Only trust the clustering when it cleanly matches the number of
            # amount columns found in the header — otherwise fall back to the
            # original x0-based logic rather than guess.
            if len(clusters) == len(numeric_cols_sorted):
                for cx, (_, ci) in zip(clusters, numeric_cols_sorted):
                    cluster_to_col[cx] = ci
                amount_clusters = clusters

    def get_amount_col_idx(x1):
        nearest = min(amount_clusters, key=lambda c: abs(c - x1))
        return cluster_to_col[nearest]

    SKIP_PHRASES = [
        'page ', 'generated', 'authorised', 'statement summary',
        'opening balance', 'closing balance', 'total debit', 'total credit',
        'eff avail', 'count of lien', ':12 pm', ':12 am', ' pm )', ' am )',
        'transaction list', 'cumulative total', 'grand total', 'brought forward',
        'end of statement', 'computer-generated', 'computer generated',
    ]
    BF_PAT = re.compile(r'\bb\s*/\s*f\b', re.IGNORECASE)  # "B/F" brought-forward marker
    # Divider lines between the transaction table and totals footer are
    # usually drawn with plain ASCII "-", but a statement generator that
    # renders dates using the Unicode minus sign (see DATE_SEP above) draws
    # its dividers with the same character — an ASCII-only class would leave
    # them unrecognized as dividers, so they'd fall through to being treated
    # as ordinary continuation rows instead of being dropped.
    DIVIDER_RE = re.compile(r'^[-=_*.\s‐-―−]+$')

    def is_divider_or_amount_only_row(row_words):
        """
        True for rows with no identifying date/narration text at all — e.g. a bare
        "B/F"-style opening-balance line (just two amounts, no label) or a plain
        "----" divider between the transaction table and a totals footer. These
        aren't real transactions, so they must NOT be merged as a "continuation"
        into a neighboring anchor row (that would silently splice their amounts
        into an unrelated transaction's WITHDRAWALS/DEPOSITS/BALANCE/PARTICULARS).
        """
        texts = [w['text'].strip() for w in row_words if w['text'].strip()]
        if not texts:
            return False
        if all(DIVIDER_RE.match(t) for t in texts):
            return True
        if all(AMOUNT_VALUE_RE.match(t) for t in texts):
            return True
        return False

    # ── Step 2: Collect all rows with global y ────────────────────────────
    all_categorized = []
    cumulative_y = 0

    for pg_idx, words in enumerate(all_page_words):
        y_rows = {}
        for w in words:
            y = round(w['top'])
            matched = next((ky for ky in y_rows if abs(ky - y) <= 5), None)
            if matched is None:
                matched = y
                y_rows[matched] = []
            y_rows[matched].append(w)

        sorted_ys = sorted(y_rows.keys())
        # Skip everything up to and including this page's own header band, if
        # it has one (every page whose header row was detected in Step 1, not
        # just data_start_page — see the comment there for why).
        if pg_idx in page_header_end_y:
            sorted_ys = [y for y in sorted_ys if y > page_header_end_y[pg_idx]]

        for y in sorted_ys:
            row_words = sorted(y_rows[y], key=lambda w: w['x0'])
            row_text = ' '.join(w['text'] for w in row_words).lower()

            if any(skip in row_text for skip in SKIP_PHRASES) or BF_PAT.search(row_text):
                continue

            # An anchor row has a date-pattern word within the date-column x-zone
            date_zone_words = [w for w in row_words if w['x0'] < date_col_x_max]
            date_zone_text = ' '.join(w['text'] for w in date_zone_words).strip()
            is_anchor = bool(DATE_PAT.match(date_zone_text))

            # Rows with no date AND no identifying text (bare "B/F"-style opening
            # balance amounts, or a plain "----" divider) aren't real transactions.
            # Drop them outright instead of letting them merge into a neighbor —
            # otherwise their amounts get spliced into an unrelated transaction.
            if not is_anchor and is_divider_or_amount_only_row(row_words):
                continue

            # Assign words to columns.
            # Special rule: the Date column (col 0) should ONLY contain the date value.
            # Any word starting beyond the date column x-zone belongs to narration (col 1+).
            row_by_col = {}
            for w in row_words:
                x = w['x0']
                if x < date_col_x_max:
                    ci = 0  # date zone
                elif cluster_to_col and AMOUNT_VALUE_RE.match(w['text'].strip()):
                    ci = get_amount_col_idx(w['x1'])  # right-aligned amount: match by x1 cluster
                else:
                    ci = get_col_idx(x)
                    if ci == 0:
                        ci = 1  # force into narration if get_col_idx returned 0 for non-date x
                row_by_col[ci] = (row_by_col.get(ci, '') + ' ' + w['text']).strip()

            all_categorized.append({
                'global_y': cumulative_y + y,
                'is_anchor': is_anchor,
                'by_col': row_by_col,
            })

        cumulative_y += page_heights[pg_idx]

    if not all_categorized:
        return None, None

    # ── Step 3: Two-pass merge ────────────────────────────────────────────
    anchor_indices = [i for i, r in enumerate(all_categorized) if r['is_anchor']]

    if not anchor_indices:
        return None, None

    transactions = [dict(all_categorized[i]['by_col']) for i in anchor_indices]
    anchor_global_ys = [all_categorized[i]['global_y'] for i in anchor_indices]
    anchor_set = set(anchor_indices)

    # Maximum y-distance (pts) to consider a row as part of an anchor transaction.
    # RBL narration wraps are ~7px away from the date row.
    # Canara bank statements can have narration lines up to ~36px away.
    # 45px safely catches all continuation lines without merging unrelated transactions.
    MAX_Y_GAP = 45

    for i, row in enumerate(all_categorized):
        if i in anchor_set:
            continue
        gy = row['global_y']
        # Find nearest anchor
        nearest_idx = min(range(len(anchor_global_ys)),
                          key=lambda k: abs(anchor_global_ys[k] - gy))
        distance = abs(anchor_global_ys[nearest_idx] - gy)
        if distance > MAX_Y_GAP:
            continue  # Too far away — skip, don't merge
        target = transactions[nearest_idx]
        for ci, val in row['by_col'].items():
            if ci in target:
                target[ci] = (target[ci] + ' ' + val).strip()
            else:
                target[ci] = val


    # ── Step 4: Convert to list-of-lists ─────────────────────────────────
    num_cols = len(col_names)
    data_rows = []
    for txn in transactions:
        out = [txn.get(i, '') for i in range(num_cols)]
        data_rows.append(out)

    return col_names, data_rows


def clean_numeric(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r'^[₹$€£]|^Rs\.?\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*\(?\s*(dr|cr)\s*\)?\.?$', '', s, flags=re.IGNORECASE)  # trailing Dr/Cr suffix, e.g. "645.00Dr" or "17.00(Cr)"
    s = s.replace(',', '')
    s = re.sub(r'\s+', '', s)
    try:
        return float(s)
    except ValueError:
        return value


# Bank statements commonly include a standalone "Statement Summary" / "Account
# Summary" box (Opening Balance, Dr/Cr Count, Total Debits/Credits, Closing
# Balance) somewhere on the page, separate from the actual transaction table.
# It's not a transaction and must never be extracted as one. Checking only the
# table's first row misses layouts where a title row precedes the summary's
# own label row (e.g. "Statement Summary as on ..." followed by
# "Opening Balance | Dr Count | ... | Closing Balance" as row 2), so scan the
# whole table's text instead of just row 0.
SUMMARY_TABLE_PHRASES = [
    'statement summary', 'account summary', 'summary as on',
    'opening balance', 'closing balance', 'total debit', 'total credit',
    'transaction total', 'grand total', 'sub total', 'dr count', 'cr count',
]


def is_summary_table(table):
    if not table:
        return False
    table_text = ' '.join(str(c).lower() for row in table for c in row if c)
    return any(p in table_text for p in SUMMARY_TABLE_PHRASES)


# Some banks' PDFs draw the table border around the trailing "TRANSACTION
# TOTAL"/"CLOSING BALANCE" summary rows too, so pdfplumber's gridline
# extraction returns them as the LAST 1-2 rows of the SAME table object as
# every real transaction above them — not as a separate standalone summary
# box. is_summary_table() (designed for a genuinely standalone box) can't
# tell these two cases apart on its own, since both contain the same marker
# phrases; skipping the whole table wholesale in the mixed case throws away
# every real transaction in it as collateral damage. Confirmed on a real
# Axis Bank statement: an entire page's worth of 23 real transactions was
# silently dropped this way, leaving only the first page's data — which
# looked like "the scanner only reads one page" but was actually this.
# DATE_SEP (not plain "[/-]") so this also recognizes dates rendered with the
# Unicode minus sign (confirmed on a real Cosmos Co-operative Bank statement
# — see DATE_SEP's own comment).
_TABLE_DATE_RE = re.compile(r'^\d{1,2}' + DATE_SEP + r'(?:\d{1,2}|[A-Za-z]{3,9})' + DATE_SEP + r'\d{2,4}$', re.IGNORECASE)


def table_has_dated_row(table):
    """True if any row's first cell looks like a real transaction date —
    meaning this table has genuine transactional content and must not be
    discarded wholesale just because a trailing summary row also matches
    is_summary_table()'s phrase check. A genuinely standalone summary box
    has no dated rows at all, so this leaves that case unaffected."""
    for row in table:
        if row and row[0] and _TABLE_DATE_RE.match(str(row[0]).strip()):
            return True
    return False


# A real header row is made of several short label cells ("Date", "Debit",
# "Particulars"). A non-tabular info box (e.g. a "Customer Details" box with
# one cell holding a whole paragraph of free text) is not, but the old
# matching below joined a row's/block's cells into one blob and substring-
# searched that blob — so a single long paragraph cell that happens to
# mention "Date of Birth" and "Total Balance" in prose satisfied the same
# "has 'date' + >=2 keyword" check as a genuine header row. Confirmed on a
# real Bank of Maharashtra statement: its "Customer Details / Branch &
# Account Details" info box (2 columns, each a multi-line paragraph) was
# misidentified as the transaction header before the real 8-column
# "Date | Type | Particulars | Cheque/Reference No | Debit | Credit |
# Balance | Channel" header ever got the chance (headers is only ever set
# once, from the FIRST row that matches) — truncating every real
# transaction row down to 2 columns, and the bogus header's second column
# ("...Total Balance...") coincidentally matching the 'balance' amount-
# keyword caused every row's Type value ("Cheque"/"Charges"/etc, non-
# numeric) to be coerced to NaN and then dropped entirely by the final
# all-NaN safety-net filter — a completely empty result with no error.
HEADER_CELL_MAX_LEN = 60


def detect_header(table, target_headers):
    def short_cells(row):
        return [str(c).strip().lower() for c in row if c and len(str(c).strip()) <= HEADER_CELL_MAX_LEN]

    for i, row in enumerate(table):
        row_cells = short_cells(row)
        block = table[i:min(i+3, len(table))]
        block_cells = [c for r in block for c in short_cells(r)]

        matches = sum(1 for t in target_headers if any(keyword_matches(c, t) for c in block_cells))
        has_date = any(keyword_matches(c, 'date') for c in block_cells)

        # Ensure the current row actually contains some of the header keywords
        # otherwise we might just be looking at the row BEFORE the header
        row_matches = sum(1 for t in target_headers if any(keyword_matches(c, t) for c in row_cells))

        if has_date and matches >= 2 and row_matches >= 1:
            col_names = [str(c).replace('\n', ' ').strip() if c else f'Column_{j}' for j, c in enumerate(row)]
            return i, col_names
    return -1, None

def find_col_offset(first_row, main_headers):
    if not first_row:
        return 0, False
    row_lower = [str(c).strip().lower() if c else '' for c in first_row]
    headers_lower = [str(h).strip().lower() for h in main_headers]
    col_diff = len(main_headers) - len(first_row)
    if col_diff <= 0:
        return 0, False
    best_offset, best_matches = 0, 0
    for offset in range(col_diff + 1):
        if offset + len(row_lower) <= len(headers_lower):
            matches = sum(1 for i, cell in enumerate(row_lower) if cell and (cell in headers_lower[i + offset] or headers_lower[i + offset] in cell))
            if matches > best_matches:
                best_matches, best_offset = matches, offset
    if best_matches >= 2:
        return best_offset, True
    return 0, False

def split_crdr_column(df):
    crdr_col = None
    amount_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'cr/dr' in col_lower:
            crdr_col = col
        elif 'amount' in col_lower and 'balance' not in col_lower and 'available' not in col_lower:
            amount_col = col

    # Some statements name this column something generic like "Type" instead
    # of "Cr/Dr" (confirmed on a real PNB statement — header literally
    # "Date Amount Type Balance Remarks") — the header text alone can't
    # reliably identify it in that case. Fall back to checking the column's
    # VALUES instead: a genuine Cr/Dr indicator column only ever contains
    # "CR"/"DR" (or "C"/"D"), so a column where the overwhelming majority of
    # non-blank values match that narrow set is safe to treat as one
    # regardless of its header name — a real "Transaction Type" column
    # (NEFT/UPI/IMPS/Cash) would never pass this check, so this can't
    # misfire and mistake an unrelated column for the direction indicator.
    if not crdr_col and amount_col:
        for col in df.columns:
            if col == amount_col:
                continue
            values = df[col].astype(str).str.strip().str.upper()
            non_blank = values[values != '']
            if non_blank.empty:
                continue
            if non_blank.isin(['CR', 'DR', 'C', 'D']).mean() >= 0.9:
                crdr_col = col
                break

    if crdr_col and amount_col:
        df['Credit'] = df.apply(lambda r: r[amount_col] if str(r[crdr_col]).upper().strip() in ('CR', 'C') else None, axis=1)
        df['Debit'] = df.apply(lambda r: r[amount_col] if str(r[crdr_col]).upper().strip() in ('DR', 'D') else None, axis=1)
        df.drop(columns=[crdr_col, amount_col], inplace=True)
    return df


CRDR_INLINE_RE = re.compile(r'\(?\s*(?:dr|cr)\s*\)?\.?\s*$', re.IGNORECASE)


def split_embedded_crdr_amount(df):
    """
    Some statements (e.g. passbook-style exports) have a single Amount column
    with the direction embedded in each cell instead of a separate Cr/Dr
    column — e.g. "17.00(Cr)" / "236.00(Dr)". Left alone, that direction is
    lost the moment the amount is converted to a plain float. Detect that
    pattern and split it into proper Debit/Credit columns before any numeric
    cleanup runs, so the sign survives.
    """
    if any('cr/dr' in str(c).lower() for c in df.columns):
        return df  # already has an explicit Cr/Dr column — split_crdr_column handles it

    for col in list(df.columns):
        col_lower = str(col).lower()
        if 'amount' not in col_lower or 'balance' in col_lower or 'available' in col_lower:
            continue

        values = df[col].astype(str)
        non_blank = values[values.str.strip() != '']
        if non_blank.empty:
            continue
        marked_ratio = non_blank.str.contains(CRDR_INLINE_RE).mean()
        if marked_ratio < 0.6:
            continue  # doesn't look like an inline-direction amount column

        def parse_amt(v):
            s = str(v).strip()
            m = re.search(r'(dr|cr)', s, flags=re.IGNORECASE)
            num = clean_numeric(s)
            if not isinstance(num, (int, float)):
                return (None, None)
            direction = m.group(1).lower() if m else None
            if direction == 'dr':
                return (num, None)
            if direction == 'cr':
                return (None, num)
            return (None, None)

        parsed = df[col].apply(parse_amt)
        df['Debit'] = parsed.apply(lambda t: t[0])
        df['Credit'] = parsed.apply(lambda t: t[1])
        df.drop(columns=[col], inplace=True)

    return df

# =================================================================
# Main parser (HYBRID MODE: Text + OCR)
# =================================================================

def parse_bank_statement(file_path, output_excel_path):
    print(f"\n--- Loading File: {file_path} ---")

    target_headers = [
        "date", "txn date", "transaction date", "value date", "posting date",
        "balance", "amount", "withdrawal", "deposit", "value date",
        "narration", "txn", "transaction", "category", "chq", "cheque", "ref", "particulars", "remarks"
    ]
    
    pages_tables = []
    is_native_pdf = False
    ext = file_path.lower().rsplit('.', 1)[-1]
    early_coord_headers, early_coord_rows = None, None

    # ── Auto-Detect Document Engine ──────────────────────────────────────
    if ext == 'pdf':
        with pdfplumber.open(file_path) as pdf:
            if len(pdf.pages) > 0:
                text = pdf.pages[0].extract_text() or ""
                # If there's selectable text, use native, fast pdfplumber
                if len(text.strip()) > 50:
                    is_native_pdf = True
                    print("[+] Native PDF detected. Using fast data extraction.")
                    ts_lines = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}
                    ts_text = {"horizontal_strategy": "text", "snap_x_tolerance": 10}

                    p0_lines = pdf.pages[0].extract_tables(table_settings=ts_lines)

                    if p0_lines and (len(p0_lines) > 5 or any(len(t) >= 3 and len(t[0]) >= 4 for t in p0_lines)):
                        print("    -> Detected explicit gridlines.")
                        ts = ts_lines
                    else:
                        # A short header LABEL (e.g. "Date") can start further
                        # RIGHT than the actual DATA values in that column do
                        # (confirmed on a real Cosmos Co-operative Bank
                        # statement: header "Date" starts at x=42, but real
                        # date values like "14-02-2023" start at x=27, left of
                        # the header-derived column boundary at x=40).
                        # get_explicit_columns() below derives column
                        # boundaries purely from header-word x-positions, so in
                        # that case pdfplumber's "explicit vertical lines"
                        # table strategy splits a value's own text mid-word
                        # across two columns (e.g. "14-02-2023" comes apart
                        # into "14-" in a phantom leading column and "02-2023"
                        # in the real Date column) — corrupting the date on
                        # EVERY row. This isn't cosmos.pdf-specific: any
                        # statement whose header label is narrower than its
                        # data column is at risk. The word-coordinate
                        # extractor doesn't have this failure mode (it anchors
                        # rows by an actual date PATTERN match, not a fixed
                        # x-boundary, and already clusters right-aligned wide
                        # amounts correctly) — try it FIRST whenever there are
                        # no gridlines, before falling back to the more
                        # fragile header-position boundary approach.
                        early_coord_headers, early_coord_rows = extract_transactions_by_word_coords(file_path, target_headers)
                        if early_coord_headers and early_coord_rows:
                            print(f"    -> No gridlines. Word-coordinate extractor found {len(early_coord_rows)} transactions.")
                            ts = None
                        else:
                            explicit_cols = get_explicit_columns(pdf.pages[0], target_headers)
                            if explicit_cols and len(explicit_cols) > 3:
                                print("    -> No gridlines. Using smart header-aligned column boundaries.")
                                ts = {
                                    "vertical_strategy": "explicit",
                                    "explicit_vertical_lines": explicit_cols,
                                    "horizontal_strategy": "text",
                                    "snap_x_tolerance": 5,
                                    "intersection_x_tolerance": 1500
                                }
                            else:
                                print("    -> No gridlines or clear headers. Using text strategy.")
                                ts = ts_text

                    if ts is not None:
                        for page in pdf.pages:
                            pages_tables.append(page.extract_tables(table_settings=ts))

    if not is_native_pdf:
        # This script has no local OCR engine of its own — a scanned/image
        # PDF (or a native PDF pdfplumber couldn't pull real text from) is
        # handed straight to Mistral OCR instead, which already does real
        # OCR as part of its normal extraction and does it reliably. An
        # earlier version of this script ran its own Tesseract-based OCR
        # pass here first, but that added a hard system dependency
        # (Tesseract-OCR + Poppler, neither pip-installable) that caused
        # real deployment failures across environments (missing on a fresh
        # VPS, missing libGL.so.1 for the GUI-linked opencv-python build,
        # missing on a local Windows dev box), and even when installed
        # correctly, running full OCR across every page of a large scanned
        # statement before ever trying Mistral could take several minutes
        # — all wasted whenever that OCR pass wasn't going to produce
        # usable output anyway, which was common. Failing fast here lets
        # pythonBankParser.ts fall through to Mistral immediately instead,
        # exactly as if this script had crashed trying to OCR it — same
        # fallback path, none of the wasted time or deployment fragility.
        print("[!] Image or Scanned PDF detected - no local OCR engine in this script. Falling back to Mistral OCR.")
        return False

    # ── Map Tables to Transactions ────────────────────────────────────────
    all_data = []
    headers = None

    # The word-coordinate extractor already ran up front for a no-gridline
    # PDF (see the "Auto-Detect Document Engine" block above) and succeeded —
    # use its result directly rather than re-deriving anything below.
    if early_coord_headers and early_coord_rows:
        headers = early_coord_headers
        all_data = early_coord_rows

    p0_tables = pages_tables[0] if pages_tables else []
    one_table_per_txn = (len(p0_tables) > 5 and len(p0_tables) > 1 and all(len(t) <= 8 for t in p0_tables[1:min(6, len(p0_tables))]))

    # Check if data appears fragmented (too many single/double-cell rows)
    # If so, use the word-coordinate extractor for accurate results
    if not (early_coord_headers and early_coord_rows) and is_native_pdf and not one_table_per_txn:
        all_flat = [row for tables in pages_tables for table in tables for row in table]
        if all_flat:
            empty_ratio = sum(1 for row in all_flat if sum(1 for c in row if c and str(c).strip()) <= 2) / len(all_flat)
            if empty_ratio > 0.35:
                print("    -> High fragmentation detected. Switching to word-coordinate extractor.")
                coord_headers, coord_rows = extract_transactions_by_word_coords(file_path, target_headers)
                if coord_headers and coord_rows:
                    headers = coord_headers
                    all_data = coord_rows
                    print(f"    -> Word-coordinate extractor found {len(coord_rows)} transactions.")

    coord_extractor_succeeded = headers is not None and bool(all_data)
    tables_to_process = [] if coord_extractor_succeeded else pages_tables  # Skip entirely if word-coord extractor already filled data

    for tables in tables_to_process:
        if one_table_per_txn:
            # ICICI-style logic
            for table in tables:
                if not table: continue
                h_idx, h_names = detect_header(table, target_headers)
                if h_idx != -1:
                    if headers is None:
                        end = min(h_idx + 3, len(table))
                        headers = merge_multirow_header(table[h_idx:end])
                    continue
                if headers is None: continue
                # No header found in this table AND it looks like a standalone
                # summary box (Opening/Closing Balance, Dr/Cr Count, etc.) rather
                # than a single-transaction table — not a real transaction.
                if is_summary_table(table): continue

                row = collapse_table_to_row(table)
                if not any(v.strip() for v in row): continue
                if len(row) < len(headers):
                    row.extend([''] * (len(headers) - len(row)))
                elif len(row) > len(headers):
                    row = row[:len(headers)]
                all_data.append(row)
        else:
            # Standard style logic
            for table in tables:
                if not table: continue
                h_idx, h_names = detect_header(table, target_headers)
                if h_idx != -1 and headers is None:
                    headers = h_names
                if headers is None: continue

                start_row = h_idx + 1 if h_idx != -1 else 0
                col_offset = 0

                if h_idx != -1:
                    while start_row < len(table):
                        sub_row = [str(c).strip().lower() for c in table[start_row] if c]
                        if not sub_row or (any(w in sub_row for w in ['date', 'no', 'description']) and not any(any(char.isdigit() for char in cell) for cell in sub_row)):
                            start_row += 1
                        else:
                            break
                else:
                    # This table has no recognized transaction header of its own.
                    # If it ALSO looks like a standalone summary box (Opening/
                    # Closing Balance, Dr/Cr Count, etc.) AND has no dated rows
                    # of its own, it isn't a continuation of the real
                    # transaction table — skip it entirely rather than risk
                    # splicing its cells in as a bogus transaction. But a table
                    # with real dated rows is a genuine continuation (e.g. a
                    # later page's table, no header repeated) even if a
                    # trailing "TRANSACTION TOTAL"/"CLOSING BALANCE" row got
                    # captured inside the SAME bordered grid — discarding the
                    # whole table in that case would silently drop every real
                    # transaction in it. The blank-date marker-row filter
                    # further down strips just the trailing summary row(s)
                    # once real per-row processing runs.
                    if is_summary_table(table) and not table_has_dated_row(table): continue
                    col_diff = len(headers) - len(table[0])
                    if col_diff < -1 or col_diff > 3: continue
                    if col_diff > 0:
                        offset, skip_row = find_col_offset(table[0], headers)
                        if not skip_row and offset == 0: continue
                        col_offset = offset
                        if skip_row: start_row = 1

                current_txn = None
                for row in table[start_row:]:
                    clean_row = [str(c).replace('\n', ' ').strip() if c else '' for c in row]
                    if not any(clean_row): continue
                    if col_offset > 0: clean_row = [''] * col_offset + clean_row

                    if len(clean_row) < len(headers):
                        clean_row.extend([''] * (len(headers) - len(clean_row)))
                    elif len(clean_row) > len(headers):
                        clean_row = clean_row[:len(headers)]

                    # A row whose date cell is either blank OR isn't a real
                    # date at all (e.g. the literal word "Total" landing in
                    # the date column — confirmed on a real Allahabad Bank
                    # statement, whose trailing "Total 51,220.76 50,411.00"
                    # summary line extracted as ['Total', '', '', '51,220.76',
                    # '50,411.00', ''], with "Total" sitting exactly where the
                    # date normally goes) is a standalone statement-summary
                    # line if its content also matches a known marker phrase,
                    # not a wrapped continuation of the row above it. Left
                    # alone, the merge logic below folds it straight into that
                    # transaction's narration — which the is_balance_marker_row
                    # filter further down then matches and deletes WHOLE,
                    # taking the genuine transaction down with it as
                    # collateral damage; or, as with the Allahabad case, the
                    # summary row survives untouched as its own bogus
                    # "transaction" with a fake ~2x-inflated Debit/Credit
                    # value. Confirmed on a real SBI statement: this exact
                    # merge mechanism silently dropped the statement's very
                    # last transaction, a real ₹95 INTEREST CREDIT. Drop the
                    # marker row outright here, before it ever reaches the
                    # merge step.
                    row_date_is_real = bool(clean_row[0]) and bool(_TABLE_DATE_RE.match(clean_row[0]))
                    if not row_date_is_real and re.search(
                        r'\bb\s*/\s*f\b|brought forward|carried forward|\bc\s*/\s*f\b|'
                        r'closing balance|opening balance|\btotal\b|'
                        r'grand total|sub\s*total',
                        ' '.join(clean_row), re.IGNORECASE,
                    ):
                        continue

                    non_empty_count = sum(1 for c in clean_row if c)
                    is_new_txn = bool(clean_row[0]) or non_empty_count >= 3

                    if is_new_txn:
                        if current_txn: all_data.append(current_txn)
                        current_txn = clean_row
                    else:
                        target = current_txn if current_txn else (all_data[-1] if all_data else None)
                        if target is not None:
                            for idx in range(min(len(clean_row), len(target))):
                                if clean_row[idx]:
                                    target[idx] = (target[idx] + ' ' + clean_row[idx]).strip()

                if current_txn:
                    all_data.append(current_txn)

    # ── Build DataFrame and Export ───────────────────────────────────────
    if not all_data:
        print("No transactional data found. Check the document format.")
        return False

    if headers is None:
        num_cols = max(len(r) for r in all_data)
        headers = [f"Column_{i + 1}" for i in range(num_cols)]

    df = pd.DataFrame(all_data, columns=headers)
    df = split_embedded_crdr_amount(df)  # must run before numeric cleanup strips the Dr/Cr direction
    # 'withdraw' (not 'withdrawal') so this also matches header variants like
    # "WITHDRAWS" that don't end in "-al" — confirmed on a real Canara
    # statement where "WITHDRAWS" fell through this check entirely, leaving
    # its values as raw comma-formatted strings (e.g. "16,000.00") instead of
    # being cleaned into a proper number. 'withdraw' is a strict prefix of
    # 'withdrawal', so this matches every case the old keyword did plus more.
    amount_keywords = ['amt', 'amount', 'balance', 'debit', 'credit', 'withdraw', 'deposit']
    for col in df.columns:
        col_lower_no_spaces = str(col).lower().replace(' ', '').replace('\n', '')
        # Bare "DR"/"CR" columns (confirmed on a real Allahabad Bank statement:
        # header literally "Post Date | Value Date | Description | DR | CR |
        # Balance", each row having exactly one of the two populated — the same
        # shape as a Withdrawal/Deposit pair, just abbreviated to 2 letters) need
        # an EXACT match, not a substring one — "dr"/"cr" as a substring check
        # would false-positive on unrelated column names like "Address" or
        # "Order No" that happen to contain those two letters.
        is_bare_dr_cr = col_lower_no_spaces in ('dr', 'cr')
        if (any_keyword_matches(col_lower_no_spaces, amount_keywords) or is_bare_dr_cr) and 'cr/dr' not in col_lower_no_spaces:
            df[col] = df[col].apply(clean_numeric)
            df[col] = pd.to_numeric(df[col], errors='coerce')

    df = split_crdr_column(df)

    # ── Clean Date column: extract only the date, prepend overflow to narration ──
    date_col = next((c for c in df.columns if str(c).lower().strip() in ('date', 'transaction date', 'txn date', 'value date') and 'value' not in str(c).lower()), None)
    narr_col = next((c for c in df.columns if any_keyword_matches(str(c).lower(), ('detail', 'narration', 'particulars', 'description', 'transaction'))), None)

    if date_col and narr_col:
        # Same numeric-or-month-name widening as DATE_PAT above, for consistency
        # on statements that reach this "standard style" path with a month-name date,
        # plus the same DATE_SEP (Unicode-minus-aware) separator class.
        DATE_RE = re.compile(r'\d{1,2}' + DATE_SEP + r'(?:\d{1,2}|[A-Za-z]{3,9})' + DATE_SEP + r'\d{2,4}', re.IGNORECASE)
        def split_date_narr(row):
            raw = str(row[date_col]).strip() if pd.notna(row[date_col]) else ''
            m = DATE_RE.search(raw)
            if m:
                date_val = m.group(0)
                overflow = (raw[:m.start()] + ' ' + raw[m.end():]).strip()
                narr_val = str(row[narr_col]).strip() if pd.notna(row[narr_col]) else ''
                if overflow:
                    narr_val = (overflow + ' ' + narr_val).strip()
                return date_val, narr_val
            return raw, str(row[narr_col]).strip() if pd.notna(row[narr_col]) else ''

        df[[date_col, narr_col]] = df.apply(
            lambda r: pd.Series(split_date_narr(r)), axis=1
        )

    # Drop non-transactional rows: bare "B/F"/"Brought Forward" opening-balance
    # lines, and rows with neither a date nor any narration text (nothing to
    # identify them as a real transaction). Applies to every extraction path
    # (OCR, ICICI-style, standard-style, word-coordinate) since it runs on the
    # final DataFrame, not inside any one extractor.
    if narr_col:
        is_bf_row = df[narr_col].astype(str).str.contains(
            r'\bb\s*/\s*f\b|brought forward', case=False, regex=True, na=False
        )
        df = df[~is_bf_row]

        # Same treatment for "Closing Balance" / "Opening Balance" carry-forward
        # markers: these are summary labels, not real transactions. This catches
        # the case where such a line rides along as its own row inside an
        # otherwise-real transaction table (is_summary_table only discards a
        # whole table when it has no real header — a single closing/opening
        # balance row embedded in a legitimate table still needs this net).
        is_balance_marker_row = df[narr_col].astype(str).str.contains(
            r'closing balance|opening balance', case=False, regex=True, na=False
        )
        df = df[~is_balance_marker_row]

    if date_col and narr_col:
        date_blank = df[date_col].isna() | (df[date_col].astype(str).str.strip() == '')
        narr_blank = df[narr_col].isna() | (df[narr_col].astype(str).str.strip() == '')
        df = df[~(date_blank & narr_blank)]

    # Filter out non-transactional junk (rows where ALL amount columns are NaN/empty)
    final_amount_cols = [
        col for col in df.columns
        if (any_keyword_matches(str(col).lower().replace(' ', '').replace('\n', ''), amount_keywords)
            or str(col).lower().replace(' ', '').replace('\n', '') in ('dr', 'cr'))
        and 'cr/dr' not in str(col).lower().replace(' ', '').replace('\n', '')
    ]
    if final_amount_cols:
        df.dropna(subset=final_amount_cols, how='all', inplace=True)

    df.replace("", pd.NA, inplace=True)
    df.dropna(how='all', inplace=True)

    df.to_excel(output_excel_path, index=False, engine='openpyxl')
    
    wb = load_workbook(output_excel_path)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    wb.save(output_excel_path)
    
    json_path = output_excel_path.rsplit('.', 1)[0] + '.json'
    df.to_json(json_path, orient='records', indent=4)
    
    print(f"[Success] Data extracted and saved to {output_excel_path} and {json_path}")
    return True

# =================================================================
# Entry point
# =================================================================

if __name__ == "__main__":
    import sys
    from datetime import datetime

    # Usage: python parse_bank_statement.py [input_file] [output_excel_path]
    # Called with explicit paths by src/lib/pythonBankParser.ts (the Next.js
    # app's subprocess wrapper). With no arguments, behaves exactly as before —
    # falls back to the original standalone default (uno.pdf) for manual runs.
    if len(sys.argv) >= 2:
        FILE_PATH = sys.argv[1]
    else:
        FILE_PATH = "bom.pdf"  # Put ANY file type here: Native PDF, Scanned PDF, .jpg, .png

    if not os.path.exists(FILE_PATH):
        print(f"File not found: {FILE_PATH}. Please provide a valid file path.")
        sys.exit(1)

    if len(sys.argv) >= 3:
        EXCEL_OUTPUT_PATH = sys.argv[2]
    else:
        file_stem = os.path.splitext(os.path.basename(FILE_PATH))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        EXCEL_OUTPUT_PATH = f"{file_stem}_{timestamp}.xlsx"

    start_time = time.time()
    try:
        success = parse_bank_statement(FILE_PATH, EXCEL_OUTPUT_PATH)
    except PermissionError:
        print(f"\n[ERROR] Cannot write to '{EXCEL_OUTPUT_PATH}'. Please close it in Excel.")
        sys.exit(1)
    else:
        elapsed = time.time() - start_time
        if success:
            print(f"Extraction completed in {elapsed:.2f} seconds.")
            sys.exit(0)
        else:
            sys.exit(1)